from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, make_response, current_app, Response
from .models import Victim, Admin, mask_email, mask_phone
from werkzeug.security import check_password_hash
from . import db
import os
import io
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from app.utils import get_ngrok_url

main = Blueprint('main', __name__)
DOWNLOAD_PIN = "a2025A"  # غير الرقم هذا للرقم اللي تبيه

@main.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        phone = request.form.get('phone')
        username = request.form.get("username")

        victim = Victim(email=email,phone=phone ,username =username)
        db.session.add(victim)
        db.session.commit()
        return redirect(url_for("main.awareness"))
    
    ngrok_url = get_ngrok_url() or "http://localhost:5000"
    return render_template("login.html", ngrok_url=ngrok_url)
@main.route('/awareness')
def awareness():
    return render_template('awareness.html')

@main.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        admin = Admin.query.filter_by(username=username).first()
        if admin and check_password_hash(admin.password, password):
            session['admin_logged_in'] = True
            return redirect(url_for("main.dashboard"))
        else:
            return render_template("admin_login.html", error="بيانات الدخول غير صحيحة")
    return render_template("admin_login.html")

@main.route("/logout")
def logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for("main.admin_login"))

@main.route("/dashboard")
def dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for("main.admin_login"))

    search_query = request.args.get('search', '').strip()
    if search_query:
        victims = Victim.query.filter(
            (Victim.email.contains(search_query)) | 
            (Victim.username.contains(search_query))
        ).all()
    else:
        victims = Victim.query.all()
    
    # عرض كلمات المرور كنجمات فقط في لوحة التحكم
    for v in victims:
        v.display_password = '*' * 8  # أو حسب عدد النجوم اللي تبيها
        v.masked_email = mask_email(v.email)
        v.masked_phone = v.phone[:3] + '****' + v.phone[-3:] if v.phone else "غير متوفر"
        v.awareness_completed = v.awareness_completed or False


    
    return render_template("dashboard.html", victims=victims, search_query=search_query)


@main.route("/login", methods=["GET"])
def phishing_login():
    email = request.args.get("e")  # استلام البريد من Gophish
    session["last_email"] = email
    ngrok_url = get_ngrok_url() or "http://localhost:5000"
    return render_template("phish_login.html", email=email, ngrok_url=ngrok_url)

@main.route("/gophish/webhook", methods=["POST"])
def gophish_webhook():
    """
    Webhook يستقبل الأحداث من Gophish ويخزن بيانات الضحايا
    مباشرة في جدول Victim الموجود.
    """
    data = request.get_json()
    if not data:
        return {"error": "No JSON received"}, 400

    # نوع الحدث (Submitted Data, Link Clicked, Email Opened)
    event_type = data.get("message", "")
    details = data.get("details", {})
    payload = details.get("payload", {})

    # محاولات متعددة للحصول على البيانات من JSON
    email = data.get("email") or payload.get("email")
    username = payload.get("username") or payload.get("name")
    phone = payload.get("phone") or payload.get("mobile")
    ip_address = details.get("ip") or request.remote_addr

    if not email:
        return {"error": "No email found in payload"}, 400

    # البحث عن الضحية في قاعدة البيانات
    victim = Victim.query.filter_by(email=email).first()
    if victim:
        # تحديث بيانات الضحية إذا موجود مسبقًا
        victim.username = username or victim.username
        victim.phone = phone or victim.phone
        victim.link_clicks += 1
        victim.phishing_count += 1
    else:
        # إضافة ضحية جديدة
        victim = Victim(
            email=email,
            username=username,
            phone=phone,
            ip_address=ip_address
        )
        db.session.add(victim)

    db.session.commit()
    return {"status": "ok", "event": event_type}, 200


@main.route("/submit_credentials", methods=["GET","POST"])
def submit_credentials():
    email = request.form.get('email')
    username = request.form.get('username')
    phone = request.form.get('phone')
    ip_address = request.remote_addr
    session["last_email"] = email

    if not email:
        flash("يرجى إدخال البريد.", "warning")
        return redirect(request.referrer)

    victim = Victim.query.filter_by(email=email).first()
    if victim:
       victim.username = username
       victim.phone = phone
       victim.link_clicks += 1
       victim.phishing_count += 1
    else:
        victim = Victim(
            email=email,
            username=username,
            phone=phone,
            ip_address=ip_address
        )
        db.session.add(victim)

    db.session.commit()

    return redirect(url_for("main.awareness"))


@main.route('/download_xlsx', methods=['POST'])
def download_xlsx():
    if not session.get('admin_logged_in'):
        return redirect(url_for("main.admin_login"))

    pin = request.form.get('pin')
    if pin != DOWNLOAD_PIN:
        return jsonify({"error": "❌ كلمة السر غير صحيحة!"}), 400

    users = Victim.query.all()

    # إنشاء ملف Excel جديد
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Victims Data"

    headers = ["Email", "Username","phone","IP Address", "Timestamp", "Link Clicked", "Phishing Count","Awareness Completed"]
    ws.append(headers)

    for user in users:
        awareness_status = 'True' if user.awareness_completed else 'False'

        ws.append([
            user.email,
            user.username,
            user.phone,
            user.ip_address,
            user.timestamp.strftime('%Y-%m-%d %H:%M:%S') if user.timestamp else "",
            user.link_clicks,
            user.phishing_count,
            awareness_status 


        ])

    # ضبط عرض الأعمدة تلقائياً
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # حفظ الملف في بايتس
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # إرجاع الملف كرد تحميل
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=victim_data.xlsx"
        }
    )

@main.route('/download_csv', methods=['POST'])
def download_csv():
    if not session.get('admin_logged_in'):
        return redirect(url_for("main.admin_login"))

    pin = request.form.get('pin')
    if pin != DOWNLOAD_PIN:
        return jsonify({"error": "❌ كلمة السر غير صحيحة!"}), 400

    users = Victim.query.all()

    output = io.StringIO()
    output.write('\ufeff')  # إضافة BOM لتشفير UTF-8
    writer = csv.writer(output)

    headers = ["Email", "Username", "phone" ,"IP Address", "Timestamp", "Link Clicked", "Phishing Count", "Awareness Completed"]
    writer.writerow(headers)

    for user in users:
        awareness_status = 'True' if user.awareness_completed else 'False'
        writer.writerow([
            user.email,
            user.username,
            user.phone,
            user.ip_address,
            user.timestamp.strftime('%Y-%m-%d %H:%M:%S') if user.timestamp else "",
            user.link_clicks,
            user.phishing_count,
            awareness_status

        ])

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=victim_data.csv"
    response.headers["Content-Type"] = "text/csv;charset=utf-8"
    return response


@main.route('/complete_awareness', methods=['POST'])
def complete_awareness():
    email = session.get("last_email")  # حفظنا البريد مسبقًا
    if email:
        victim = Victim.query.filter_by(email=email).first()
        if victim:
            victim.awareness_completed = True
            db.session.commit()

    return '', 204  # استجابة فارغة ناجحة
